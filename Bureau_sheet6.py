def bureau_s6(st_day,end_day,out_xls,user,pwd,ip,sid,oname):
    import afmod
    import afmod1
    import binascii
    from openpyxl import load_workbook

#--------------------------------------------------------
    def add_sheet(x,on_sheet2,tmp_rm01,csk06,moy17,rm101_1):
        on_sheet2.cell(row=x, column=1, value=tmp_rm01)
        on_sheet2.cell(row=x, column=2, value=csk06)
        on_sheet2.cell(row=x, column=3, value=moy17)
        on_sheet2.cell(row=x, column=5, value=rm101_1)

    # --------------------------------------------------------
    other_citys = {'B':'臺中市','D':'臺南市','E':'高雄市','F':'新北市','H':'桃園市','C':'基隆市','G':'宜蘭縣','I':'嘉義市',
             'J':'新竹縣','K':'苗栗縣','M':'南投縣','N':'彰化縣','O':'新竹市','P':'雲林縣','Q':'嘉義縣','T':'屏東縣',
             'U':'花蓮縣','V':'臺東縣','W':'金門縣','X':'澎湖縣', 'Z':'連江縣'  }
    citys = {'AA':'古亭所','AB':'建成所','AC':'中山所','AD':'松山所','AE':'士林所','AF':'大安所' }
    for city, cityname in citys.items():
        if city == oname:
            city_name = cityname
    tmp_name = "統計期間:" + st_day + '起至' + end_day + '止' +'   所別：'+city_name
    wb = load_workbook(filename=out_xls)
    wb.create_sheet('他縣市收取本市登記費報表', index=7)
    on_sheet2 = wb['他縣市收取本市登記費報表']
    on_sheet2.sheet_format.defaultColWidth = 25
    on_sheet2.merge_cells('A1:C1')
    on_sheet2.merge_cells('A2:C2')
    on_sheet2.cell(row=1, column=1, value="跨直轄市、縣(市)收辦土地登記案件他縣市收取本市登記費報表")
    on_sheet2.cell(row=2, column=1, value=tmp_name)
    add_sheet(3, on_sheet2, '跨縣市辦理轄區案號','登記原因','登記費','')
    x = 3
    str_sql99N = """
        select rm01,rm02,rm03,rm09,rm101_1,rawtohex(utl_raw.cast_to_raw(sk06.kcnt)) as skcnt06,rawtohex(utl_raw.cast_to_raw(sk02.kcnt)) as skcnt02
          from scrsms ,(select kcde_2,kcnt from srkeyn where kcde_1='06') sk06 ,(select kcde_2,kcnt from srkeyn where kcde_1='04') sk02
          where  (rm101_1 <>'A') AND sk06.kcde_2 = rm09 AND  sk02.kcde_2 = rm02 and rm07_1 >=
          """
    str_sql = str_sql99N + "\'" + st_day +"\'"+" AND rm07_1<= "+ "\'" + end_day + "\'" + ' order by rm101_1'
    #print(str_sql)
    sum_moy = 0;
    data = afmod.con_ora(str_sql, user, pwd, ip, sid)  # 連結oracle取得資料
#-----------------------------------------------------------------------------------
    for i in range(0, len(data)):
        rm01, rm02, rm03, rm09,rm101_1, sk06, sk02 = data[i]
        csk06 = binascii.unhexlify(sk06).decode('BIG5')
        csk02 = binascii.unhexlify(sk02).decode('BIG5')
        str_sql83 = '';moy17 =0;tmp_rm01 = ''; st17_2_N =0; st17_2_m=0
        if rm09 == '83':
            tmp_rm01 = rm01 + '年' + csk02 + '字第' + rm03 + '號'
            str_sql83 = "select ts_type,tc17_2 from wrtogh where ts03 ="
            str_sql83 = str_sql83 + "\'" + rm01 +"\'"+' and ts04_1 = '+ "\'" + rm02 +"\'" +' and ts04_2 = ' +"\'" + rm03 +"\'"
            str_sql83 = str_sql83 + ' and ts_type =' +"\'" + 'A' + "\'"
            data1 = afmod.con_ora(str_sql83, user, pwd, ip, sid)
            for i in range(0, len(data1)):
                t_type,st17_2 = data1[i]
                afmod1.set_rm13(t_type, st17_2)
                moy17 = int(st17_2) / 1000
            x +=1
            add_sheet(x, on_sheet2, tmp_rm01, csk06, moy17,rm101_1)
            sum_moy = sum_moy + moy17

        elif (rm09 == '85') or (rm09 =='BS'):
            tmp_rm01 = rm01 + '年' + csk02 + '字第' + rm03 + '號'
            str_sql83 = "select ts_type,tc17_2 from wrtogh where ts03 ="
            str_sql83 = str_sql83 + "\'" + rm01 + "\'" + ' and ts04_1 = ' + "\'" + rm02 + "\'" + ' and ts04_2 = ' + "\'" + rm03 + "\'"
            str_sql83 = str_sql83 + ' and ts_type in (' + "\'" + 'M' + "\'" +','+ "\'" + 'N' + "\'" +')'
            #print(str_sql83)
            data1 = afmod.con_ora(str_sql83, user, pwd, ip, sid)
            for i in range(0, len(data1)):
                t_type ,st17_2 = data1[i]
                t_type, st17_2 = afmod1.set_rm13(t_type, st17_2)
                if t_type =='M':
                    st17_2_m = int(st17_2)
                else:
                    st17_2_N = int(st17_2)
            if st17_2_N > st17_2_m :
                moy17 = (st17_2_N - st17_2_m) / 1000
            else:
                moy17 = 0
            x += 1
            sum_moy = sum_moy + moy17
            add_sheet(x, on_sheet2, tmp_rm01, csk06, moy17,rm101_1)
            # print('BS=',st17_2_N, st17_2_m,moy17)
    del binascii
    x += 3
    on_sheet2.cell(row=x, column=1, value="總計：")
    on_sheet2.cell(row=x, column=3, value=sum_moy)
    # ----------------------------------------------------------
    rm101_count = x +3
    on_sheet2.cell(row=rm101_count, column=1, value="他縣市收取本市登記費報表-以縣市統計")
    rm101_count += 1
    add_sheet(rm101_count, on_sheet2, '縣市別', '登記費總計', '', '')
    ch_name = '!';sum_city_moy = 0
    for j in range(4, x -1):
        col1_str = on_sheet2.cell(row=j, column=3).value
        col2_str = on_sheet2.cell(row=j, column=5).value
        if j == 4 :
            chname = col2_str
            sum_city_moy = col1_str
        else:
          if (chname == col2_str):
              sum_city_moy = sum_city_moy + int(col1_str)
          else:
              rm101_count += 1
              for city1, cityname1 in other_citys.items():
                  if city1 == chname:
                      city_name1 = cityname1
              add_sheet(rm101_count, on_sheet2, city_name1, sum_city_moy, '', '')
              chname = col2_str
              col1_str,col1_str = afmod1.set_rm13(col1_str,col1_str)
              sum_city_moy = int(col1_str)
       # ----------------------------------------------------------
    for j in range(4, x -1):
        on_sheet2.cell(row=j, column=5).value = ''
    wb.save(filename=out_xls)