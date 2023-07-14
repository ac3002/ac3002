def Bureau_s3(st_day,end_day,out_xls,user,pwd,ip,sid,sample_xls):
    import afmod
    import afmod1
    from openpyxl import load_workbook
    global i
#--------------------------------------------------------
    str_sql99N = """
        select DISTINCT rm101_1,rm09,rawtohex(utl_raw.cast_to_raw(kcnt)) as skcnt,Count(*) as cnt ,Sum(rm13) as srm13 ,Sum(rm16) as srm16
          from scrsms ,(select kcde_2,kcnt from srkeyn where kcde_1='06')
          where kcde_2 = rm09 AND (rm101_1 <>'A') and rm07_1 >=  """
    str_sql = str_sql99N + "\'" + st_day +"\'"+" AND rm07_1<= "+ "\'" + end_day +"\'"+" group  by rm101_1,rm09,kcnt "
#print(str_sql)
    try:
        data = afmod.con_ora(str_sql,user,pwd,ip,sid) #連結oracle取得資料
    except:
        input("-------地政資料庫連結失敗,輸入任意鍵以結束---------")
        exit()
    citys = {'B':'臺中市','D':'臺南市','E':'高雄市','F':'新北市','H':'桃園市','C':'基隆市','G':'宜蘭縣','I':'嘉義市','J':'新竹縣','K':'苗栗縣',
         'M':'南投縣','N':'彰化縣','O':'新竹市','P':'雲林縣','Q':'嘉義縣','T':'屏東縣','U':'花蓮縣','V':'臺東縣','W':'金門縣','X':'澎湖縣', 'Z':'連江縣' }
    seet_column_name = ("縣市",'登記原因','件數','筆數','棟數')
    x =0;R45 ='9';rdata =[]
 #   out_xls = prtoutfile_path+out_xls+offname +'.xlsx'
    wb = load_workbook(filename=sample_xls)
    wb.save(filename=out_xls)
    wb = load_workbook(filename=out_xls)
    wb.create_sheet('他縣市承辦本所案件資料',index=3)
    on_sheet3 = wb['他縣市承辦本所案件資料']
    on_sheet3.sheet_format.defaultColWidth = 20
    on_sheet3.cell(row=1, column=3,value="他縣市承辦本所案件資料")
    for column in seet_column_name:
        x += 1
        on_sheet3.cell(row=2, column= x, value=column)
    import binascii
    cnt1 = 3;cnt_tot =0;sum_rm13 =0;sum_rm16=0;
    for i in range(0,len(data)):
        rdata.clear()
        rm101_1, rm09,kcnt,cnt,srm13,srm16 = data[i]
        crm09 = binascii.unhexlify(kcnt).decode('BIG5')
        rm13 ,rm16 = afmod1.set_rm13(srm13, srm16)
        if R45 != rm101_1 :
            if (cnt_tot + sum_rm13 + sum_rm16) > 0:
                on_sheet3.cell(row=cnt1 + i, column=1, value='合計：')
                on_sheet3.cell(row=cnt1 + i, column=3, value=cnt_tot)
                on_sheet3.cell(row=cnt1 + i, column=4, value=sum_rm13)
                on_sheet3.cell(row=cnt1 + i, column=5, value=sum_rm16)
                cnt_tot = 0
                sum_rm13 = 0
                sum_rm16 = 0
                cnt1 += 1
        R45 = rm101_1

        for city, cityname in citys.items():
            if city == rm101_1:
                city_name = cityname
        x = 0
        rdata.append(city_name)
        rdata.append(crm09)
        rdata.append(cnt)
        rdata.append(srm13)
        rdata.append(srm16)
        cnt_tot = cnt_tot + cnt
        srm13,srm16 = afmod1.set_rm13(srm13,srm16)
        sum_rm13 = sum_rm13 +srm13
        sum_rm16 = sum_rm16 +srm16
        for column in rdata:
            x += 1
            on_sheet3.cell(row = cnt1 +i, column=x, value=column)
    on_sheet3.cell(row=cnt1 + (i +1), column=1, value='合計：')
    on_sheet3.cell(row=cnt1 + (i + 1), column=3, value=cnt_tot)
    on_sheet3.cell(row=cnt1 + (i + 1), column=4, value=sum_rm13)
    on_sheet3.cell(row=cnt1 + (i + 1), column=5, value=sum_rm16)
#------------------------------------------------------
    wb.save(filename=out_xls)

